import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
import re
import tkinter as tk
from tkinter import messagebox, filedialog
import subprocess

# ===== ファイル選択ダイアログ =====
input_file = filedialog.askopenfilename(title="OCS発送依頼書を選択", filetypes=[("Excel files", "*.xlsx")])
if not input_file:
    raise ValueError("OCS発送依頼書が選択されていません。")

template_file = filedialog.askopenfilename(title="テンプレートファイルを選択", filetypes=[("Excel files", "*.xlsx")])
if not template_file:
    raise ValueError("テンプレートファイルが選択されていません。")

#signature_image = filedialog.askopenfilename(title="署名画像を選択", filetypes=[("PNG files", "*.png")])

# ===== 日付抽出 & フォルダ作成 =====
match = re.search(r'(\d{6})', input_file)
if not match:
    raise ValueError("ファイル名から日付（6桁）が抽出できません。")
date_suffix = match.group(1)
output_folder = f"OCS_invoice_{date_suffix}"
os.makedirs(output_folder, exist_ok=True)

# ===== データ読み込み =====
xls = pd.ExcelFile(input_file)
sheet_name = xls.sheet_names[0]
df = xls.parse(sheet_name)
df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
current_index = 0

# ===== データ出力処理 =====
def export_current_row():
    global current_index
    if current_index >= len(df):
        messagebox.showinfo("完了", "すべてのInvoiceを出力しました。")
        subprocess.run(["open", output_folder])  # Finderでフォルダを開く（Mac）
        return

    row = df.iloc[current_index]
    wb = load_workbook(template_file)
    ws = wb.active

    invoice_name = f"INV_{date_suffix}_{current_index+1:03}"
    ws["D3"] = invoice_name

    ws["A7"] = str(row.get("Clinic Name", ""))
    住所 = row.get("Address", "")
    医師名 = row.get("Doctor's Name", "")
    電話番号 = row.get("TEL", "")
    value_a8 = f"{住所} {電話番号}\nDr.{医師名} +81 90 9302 0682"
    ws["A8"] = value_a8

    数量列名 = "SBC Eye Booster\n（発注単位：箱 ）\n1箱 20個"
    ws["E18"] = str(row.get(数量列名, ""))

    #if os.path.exists(signature_image):
    #    img = ExcelImage(signature_image)
    #    img.anchor = "E25"
    #    ws.add_image(img)

    output_filename = f"INV_{date_suffix}_{current_index+1:03}.xlsx"
    output_path = os.path.join(output_folder, output_filename)
    wb.save(output_path)

    messagebox.showinfo("保存完了", f"{output_filename} を保存しました。")
    current_index += 1
    update_display()

# ===== 表示更新処理 =====
def update_display():
    if current_index < len(df):
        row = df.iloc[current_index]
        display_rename = {
            "Clinic Name": "Clinic Name",
            "Doctor's Name": "Doctor Name",
            "医療法人": "Medical Corporation",
            "住所": "Address (JP)",
            "TEL": "Phone",
            "SBC Eye Booster\n（発注単位：箱 ）\n1箱 20個": "Order Quantity"
        }
        text.delete("1.0", tk.END)
        for col, val in row.items():
            label = display_rename.get(col, col)
            line = f"{label.ljust(20)}: {str(val)}\n"
            text.insert(tk.END, line)
    else:
        text.delete("1.0", tk.END)
        text.insert(tk.END, "すべてのデータを処理しました。")

# ===== GUI構築 =====
root = tk.Tk()
root.title("OCS Invoice 出力ツール")
text = tk.Text(root, width=100, height=30, font=("Courier New", 12))
text.pack()
button = tk.Button(root, text="このデータでInvoice作成", command=export_current_row)
button.pack()
update_display()
root.mainloop()
